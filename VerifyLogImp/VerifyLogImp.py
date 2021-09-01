import os
import time
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def dowriteexcleTimes(ws, nrow, strline, curstamp, sectime):
    strout = strline.split(":[")
    index = len(strout)
    #print(index)
    if index == 1:
        ws.cell(row=nrow, column=3).value = strline
    else:
        ws.cell(row=nrow, column=1).value = int(strout[0])
        if False == strout[1].isspace():
            strout2 = strout[1].split("]:", 1)
            if index > 1:
                ws.cell(row=nrow, column=2).value = int(strout2[0])
                ws.cell(row=nrow, column=3).value = strout2[1]
                if sectime != 0:
                    gaptime = int(strout2[0]) - curstamp
                    curticks = sectime + gaptime /1000
                    tmstruct = time.localtime(curticks)
                    ws.cell(row=nrow, column=4).value = time.strftime("%Y-%m-%d %H:%M:%S",tmstruct) 
    return
def dowriteexcle(ws, nrow, strline):
    strout = strline.split(":[")
    index = len(strout)
    #print(index)
    if index == 1:
        ws.cell(row=nrow, column=3).value = strline
    else:
        ws.cell(row=nrow, column=1).value = int(strout[0])
        if False == strout[1].isspace():
            strout2 = strout[1].split("]:", 1)
            if index > 1:
                ws.cell(row=nrow, column=2).value = int(strout2[0])
                ws.cell(row=nrow, column=3).value = strout2[1]
    return
def main():
    def selectlogfile():
        srcfilename = text1.get()
        print(srcfilename)
        inlogfilename = filedialog.askopenfilenames(title='选择Log文件', filetypes=[('All Files', '*')])
        print(inlogfilename)
        if len(inlogfilename) > 0:
            text1.delete(0, END)
            text1.insert(INSERT,inlogfilename)

    def selectlogfileout1():
        srcfilename = textout1.get()
        sfname = filedialog.askopenfilename(title='选择Log文件', filetypes=[('All Files', '*')])
        print(sfname)
        if len(sfname)>0:
            textout1.delete(0, END)
            textout1.insert(INSERT,sfname)
    
    def selectlogfileout2():
        srcfilename = textout2.get()
        sfname = filedialog.askopenfilename(title='选择excel文件', filetypes=[('Excel', '*.xlsx'),('All Files', '*')])
        print(sfname)
        if len(sfname)>0:
            textout2.delete(0, END)
            textout2.insert(INSERT,sfname)

    def closeThisWindow():
        root.destroy()

    def clearWording():
        text2.delete(0, END)
        text3.delete(0, END)
        text4.delete(0, END)
        text5.delete(0, END)
        text6.delete(0, END)
        text7.delete(0, END)
        text8.delete(0, END)
        text9.delete(0, END)
        text10.delete(0, END)
        text11.delete(0, END)
    def getselectlist(sltList):
        if len(text2.get()) != 0:
            sltList.append(text2.get())
        if len(text3.get()) != 0:
            sltList.append(text3.get())
        if len(text4.get()) != 0:
            sltList.append(text4.get())
        if len(text5.get()) != 0:
            sltList.append(text5.get())
        if len(text6.get()) != 0:
            sltList.append(text6.get())
        if len(text7.get()) != 0:
            sltList.append(text7.get())
        if len(text8.get()) != 0:
            sltList.append(text8.get())
        if len(text9.get()) != 0:
            sltList.append(text9.get())
        if len(text10.get()) != 0:
            sltList.append(text10.get())
        if len(text11.get()) != 0:
            sltList.append(text11.get())
        num = len(sltList)
        for i in range(num):
            print(sltList[i])

    def doProcessToLogFile():
        # check if have in and out file
        if len(text1.get()) == 0:
            tkinter.messagebox.showinfo('提示','请选择要处理log文件')
            return
        if len(textout1.get()) == 0:
            tkinter.messagebox.showinfo('提示','请选择输出log文件')
            return

        # get all in file directory and check if this file exist
        strpara = text1.get()      
        sltFileList = strpara.split()
        filenum = len(sltFileList)
        for i in range(filenum):
            if os.path.exists(sltFileList[i]):
                print('文件'+sltFileList[i]+'建立成功')
            else:
                tkinter.messagebox.showinfo('提示','选择处理的log文件:'+sltFileList[i]+'不存在')
                return
        
        filew = open(textout1.get(), 'w+')

         # get select chars
        myList = []
        getselectlist(myList)
        listnum = len(myList)

         # 筛选所有log files里的log message并写到output file
        for i in range(filenum):
            filer = open(sltFileList[i], 'r')
            filew.write(sltFileList[i])
            filew.write('\r\n')
            rline = filer.readline()
            while rline:
                for i in range(listnum):
                    if len(myList[i]) !=0 and (myList[i] in rline):
                        filew.write(rline)
                        break
                rline = filer.readline()
            filer.close()
        filew.close()
    def doProcessToExcleFileAddTimeStamp():
        ticks = time.time()
        print("current time:", ticks)
        if len(text1.get()) == 0:
            tkinter.messagebox.showinfo('提示','请选择要处理log文件')
            return
        if len(textout2.get()) == 0:
            tkinter.messagebox.showinfo('提示','请选择输出excel文件')
            return
        
        # get all in file directory and check if this file exist
        strpara = text1.get()      
        sltFileList = strpara.split()
        filenum = len(sltFileList)
        for i in range(filenum):
            if os.path.exists(sltFileList[i]):
                print('文件'+sltFileList[i]+'建立成功')
            else:
                tkinter.messagebox.showinfo('提示','选择处理的log文件:'+sltFileList[i]+'不存在')
                return
        #打开一个workbook
        print(textout2.get())

        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        strfile = "Log file:" + text1.get()
        #获取当前活跃的worksheet,默认就是第一个worksheet
        ws = wb.active
        ws.append([strfile])
        ws.append(["LineNo.","TimeStamp(msec)","LogMessage"])
        row = 3

        myList = []
        RTCList = []
        getselectlist(myList)
        listnum = len(myList)
        for j in range(filenum):
            filer = open(sltFileList[j], 'r')
            rline = filer.readline()
            while rline:
                if ("RTC:" in rline):
                    RTCList = rline.split()
                    print(RTCList)
                    break
                rline = filer.readline()
            filer.close()
        if(len(RTCList) > 1):
            RTCtimestr = RTCList[1]+' '
            RTCtimestr = RTCtimestr + RTCList[3]
            print(RTCtimestr)
            structTimes = time.strptime(RTCtimestr,'%Y-%m-%d %H:%M:%S')
            print(structTimes)
            sectime = time.mktime(structTimes)
            print(sectime)

            curStampList = RTCList[0].split(":")
            print(curStampList)
            print(len(curStampList[1]))
            curstamp = curStampList[1][1:len(curStampList[1])-1]
            print(curstamp)
        else:
            curstamp = '1'
            sectime = 0
        for i in range(filenum):
            filer = open(sltFileList[i], 'r')
            rline = filer.readline()
            while rline:
                for i in range(listnum):
                    if len(myList[i]) !=0 and (myList[i] in rline):
                        dowriteexcleTimes(ws, row, rline, int(curstamp), sectime)
                        row = row + 1
                        break
                rline = filer.readline()
            filer.close()
        #保存
        wb.save(filename=textout2.get())

    def doProcessToExcleFile():
        if len(text1.get()) == 0:
            tkinter.messagebox.showinfo('提示','请选择要处理log文件')
            return
        if len(textout2.get()) == 0:
            tkinter.messagebox.showinfo('提示','请选择输出excel文件')
            return
        
        # get all in file directory and check if this file exist
        strpara = text1.get()      
        sltFileList = strpara.split()
        filenum = len(sltFileList)
        for i in range(filenum):
            if os.path.exists(sltFileList[i]):
                print('文件'+sltFileList[i]+'建立成功')
            else:
                tkinter.messagebox.showinfo('提示','选择处理的log文件:'+sltFileList[i]+'不存在')
                return
        #打开一个workbook
        print(textout2.get())

        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        strfile = "Log file:" + text1.get()
        #获取当前活跃的worksheet,默认就是第一个worksheet
        ws = wb.active
        ws.append([strfile])
        ws.append(["LineNo.","Time(msec)","LogMessage"])
        row = 3

        myList = []
        RTCList = []
        getselectlist(myList)
        listnum = len(myList)
        for i in range(filenum):
            filer = open(sltFileList[i], 'r')
            rline = filer.readline()
            while rline:
                if ("RTC:" in rline):
                    RTCList = rline.split()
                    print(RTCList)
                    break
                rline = filer.readline()
            filer.close()

        for i in range(filenum):
            filer = open(sltFileList[i], 'r')
            rline = filer.readline()
            while rline:
                for i in range(listnum):
                    if len(myList[i]) !=0 and (myList[i] in rline):
                        dowriteexcle(ws, row, rline)
                        row = row + 1
                        break
                rline = filer.readline()
            filer.close()
        #保存
        wb.save(filename=textout2.get())
    #初始化
    root=Tk()

    #设置窗体标题
    root.title('VerifyLog@V0.01')

    #设置窗口大小和位置
    root.geometry('600x600')


    label1=Label(root,text='请选择log文件:')
    text1=Entry(root,bg='white',width=46)
    button1=Button(root,text='浏览...',width=8,command=selectlogfile)

    label2=Label(root,text='筛选词条1:')
    text2=Entry(root,bg='white',width=46)
    label3=Label(root,text='筛选词条2:')
    text3=Entry(root,bg='white',width=46)
    label4=Label(root,text='筛选词条3:')
    text4=Entry(root,bg='white',width=46)
    label5=Label(root,text='筛选词条4:')
    text5=Entry(root,bg='white',width=46)
    label6=Label(root,text='筛选词条5:')
    text6=Entry(root,bg='white',width=46)

    label7=Label(root,text='筛选词条6:')
    text7=Entry(root,bg='white',width=46)
    label8=Label(root,text='筛选词条7:')
    text8=Entry(root,bg='white',width=46)
    label9=Label(root,text='筛选词条8:')
    text9=Entry(root,bg='white',width=46)
    label10=Label(root,text='筛选词条9:')
    text10=Entry(root,bg='white',width=46)
    label11=Label(root,text='筛选词条10:')
    text11=Entry(root,bg='white',width=46)

   
    labelout1=Label(root,text='输出到log文件:')
    textout1=Entry(root,bg='white',width=46)
    buttonout1=Button(root,text='浏览...',width=8,command=selectlogfileout1)
    
    labelout2=Label(root,text='输出到excel文件:')
    textout2=Entry(root,bg='white',width=46)
    buttonout2=Button(root,text='浏览...',width=8,command=selectlogfileout2)
    button2=Button(root,text='处理到log文件',width=12,command=doProcessToLogFile)
    button3=Button(root,text='退出',width=8,command=closeThisWindow)
    button4=Button(root,text='处理到excel文件',width=12,command=doProcessToExcleFileAddTimeStamp)
    buttonclr =Button(root,text='清楚所有词条',width=12,command=clearWording)

    label1.pack()
    text1.pack()
    button1.pack()
    label2.pack()
    text2.pack()
    label3.pack()
    text3.pack()
    label4.pack()
    text4.pack()
    label5.pack()
    text5.pack()
    label6.pack()
    text6.pack()

    label7.pack()
    text7.pack()
    label8.pack()
    text8.pack()
    label9.pack()
    text9.pack()
    label10.pack()
    text10.pack()
    label11.pack()
    text11.pack()
   
    labelout1.pack()
    textout1.pack()
    buttonout1.pack()
    labelout2.pack()
    textout2.pack()
    buttonout2.pack()
    button2.pack()
    button3.pack() 
    button4.pack() 
    buttonclr.pack()

    label1.place(x=30,y=30)
    text1.place(x=150,y=30)
    button1.place(x=490,y=26)
    
    label2.place(x=30,y=60)
    text2.place(x=150,y=60)
    label3.place(x=30,y=90)
    text3.place(x=150,y=90)
    label4.place(x=30,y=120)
    text4.place(x=150,y=120)
    label5.place(x=30,y=150)
    text5.place(x=150,y=150)
    label6.place(x=30,y=180)
    text6.place(x=150,y=180)

    label7.place(x=30,y=210)
    text7.place(x=150,y=210)
    label8.place(x=30,y=240)
    text8.place(x=150,y=240)
    label9.place(x=30,y=270)
    text9.place(x=150,y=270)
    label10.place(x=30,y=300)
    text10.place(x=150,y=300)
    label11.place(x=30,y=330)
    text11.place(x=150,y=330)
    
    labelout1.place(x=30,y=360)
    textout1.place(x=150,y=360)
    buttonout1.place(x=490,y=358)
    labelout2.place(x=30,y=390)
    textout2.place(x=150,y=390)
    buttonout2.place(x=490,y=388)
    
    button4.place(x=250,y=420)
    button3.place(x=380,y=420)
    button2.place(x=120,y=420)
    buttonclr.place(x=490,y=330)
    root.mainloop() 

 
if __name__=="__main__":
    main()