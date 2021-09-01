#code size statics for simva.out.elf
import os
import openpyxl

run_step = 2 #1:only make map file; 2:only make xsl file; 3:only make search&match result; 4:make all
source_file = "source\\mfp_zolox_v106.axf"
source_file2 = "source\\mfp_v014.axf"

temp_file = "source\\codeSize.map"
temp_file2 = "source\\codeSize2.map"
conv_command = "C:\\CodeSourcery\\SourceryLite\\bin\\arm-none-eabi-nm.exe -S -l " + source_file + " > " + temp_file
conv_command2 = "C:\\CodeSourcery\\SourceryLite\\bin\\arm-none-eabi-nm.exe -S -l " + source_file2 + " > " + temp_file2
Statics_sheet = "statics"
Statics_sheet2 = "statics2"
Statics_result_file = "codeSizeStatics.xlsx"

ModuleDict = {"BSP":'BSP', "DPS":'DPS', "DENS":'net', "DENS_M":'net', "EMUL":'EMUL', "FAX":'FAX', "FAXCORE":'FAXCORE', "MPRINT":'MPRINT',"SYS":'SYS',  \
    "inferno":'inferno', "ipssky":'ipssky', "VOS":'VOS'}

ModuleSize = {"BSP":0, "DPS":0, "DENS":0, "DENS_M":0, "EMUL":0, "FAX":0, "FAXCORE":0, "MPRINT":0,"SYS":0,  \
    "inferno":0, "ipssky":0, "VOS":0, "other":0, "noName":0}

def mk_lineList(lineList):
    #for so lib
    if(lineList[0] == 'U'): #.so
        if '@@' in lineList[1]:
            name, soLib = lineList[1].split('@@')
        elif '@' in lineList[1]:
            name, soLib = lineList[1].split('@')
        else:
            name, soLib = lineList[1], 'NONE'
        lineList[0] = 'so'
        lineList[1] = 'so'
        lineList.append('text')
        lineList.append(name)
        lineList.append(soLib)
        return 0
    
    #for no name
    cnt_lineList = len(lineList)
    if(cnt_lineList < 4):
        return 1

    #Type
    if((lineList[2] == 't') or (lineList[2] == 'T')): 
        lineList[2] = 'text'
    elif((lineList[2] == 'b') or (lineList[2] == 'B')):
        lineList[2] = 'bss'
    elif((lineList[2] == 'd') or (lineList[2] == 'D')):    
        lineList[2] = 'data'
    elif((lineList[2] == 'r') or (lineList[2] == 'R')):    
        lineList[2] = 'rdata'
    else:
        return 1

    #size
    bb = int(lineList[1], 16)
    lineList[1] = bb

    #remove unvalid symbl name:
    if "__func__" in lineList[3]:
        return 1
    if "__FUNCTION__" in lineList[3]:
        return 1

    #no fileName
    if(cnt_lineList == 4):
        ModuleSize['noName'] += bb
        return 0
    
    #Module
    whole_fileName = lineList[4]
    lineList[4] = 'other'
    for keys in ModuleDict.keys():
        if keys in whole_fileName:
            lineList[4] = ModuleDict[keys]
            ModuleSize[keys] += bb
            break
    if(lineList[4] == 'other'):
        ModuleSize['other'] += bb
        
    #FileName
    filePathList = whole_fileName.split('/')
    cnt = len(filePathList)
    fileName = filePathList[cnt-1]
    fileName2, lines = fileName.split(':')
    lineList.append(fileName2)
    return 0

#---------------------------------------------------------------------------
#--make map file
if((run_step == 1) or (run_step > 3)): 
    print(conv_command+'...')
    os.system(conv_command)
    print('done!\r\n'+conv_command2+'...')
    os.system(conv_command2)
    print('done!\r\n')

#--make xsl file
if((run_step == 2) or (run_step > 3)): 
    print("make xsl file\r\n")
    wb = openpyxl.Workbook()

    #####################for source_file
    ws = wb.active
    ws.title = Statics_sheet
    ws.append(["Address", "Size", "Type", "Name", "Module", "FileName"])

    line_cnt = 0
    fd_temp = open(temp_file, 'r')
    for line in fd_temp.readlines():
        line_ls = line.split()
        if(mk_lineList(line_ls) == 0):
            ws.append(line_ls)
            line_cnt += 1

    print("line_cnt=", line_cnt)
    fd_temp.close()
    ws.append([" ", " ", " ", " ", " ", " "])
    ws.append(list(ModuleSize.keys()))
    ws.append(list(ModuleSize.values()))

    #####################for source_file2
    ws2 = wb.create_sheet(Statics_sheet2)
    ws2.title = Statics_sheet2
    ws2.append(["Address", "Size", "Type", "Name", "Module", "FileName", "size on sheet1", "changed"])

    line_cnt = 0
    fd_temp = open(temp_file2, 'r')
    for line in fd_temp.readlines():
        line_ls = line.split()
        if(mk_lineList(line_ls) == 0):
            ws2.append(line_ls)
            line_cnt += 1

    print("line_cnt=", line_cnt)
    fd_temp.close()
    ws2.append([" ", " ", " ", " ", " ", " "])
    ws2.append(list(ModuleSize.keys()))
    ws2.append(list(ModuleSize.values()))

    wb.save(Statics_result_file)

#--make search&match result
if((run_step == 3) or (run_step > 3)): 
    print("get name&size\r\n")
    #--open sheet
    wb = openpyxl.load_workbook(Statics_result_file, data_only=True)
    ws = wb[Statics_sheet]
    ws2 = wb[Statics_sheet2]

    #--get size&name list
    ws_size_list = list()
    for col in ws['B']:
        ws_size_list.append(col.value)

    ws_name_list = list()
    for col in ws['D']:
        ws_name_list.append(col.value)    

    rowCnt = ws2.max_row-4
    ws2_name_list = list()
    for cnt, col in enumerate(ws2['D']):
        ws2_name_list.append(col.value) 
        if(cnt > rowCnt):
            break

    ws2_size_list = list()
    for cnt, col in enumerate(ws2['B']):
        ws2_size_list.append(col.value) 
        if(cnt > rowCnt):
            break

    print("search match name\r\n")
    #--search match name
    ws2_matchsize_list = list()
    for cnt, name in enumerate(ws2_name_list):
        matchFlag = 0
        for index, name1 in enumerate(ws_name_list):
            if(name == name1):
                ws2_matchsize_list.append(ws_size_list[index])
                matchFlag = 1
                break
        if(matchFlag == 0):
            ws2_matchsize_list.append(0)

    print("write match size\r\n")
    #--write match size
    for index, row in enumerate(ws2.rows):
        if(index == 0):
            continue
        row[6].value =  ws2_matchsize_list[index]   
        row[7].value =  int(ws2_size_list[index] - int(ws2_matchsize_list[index]))   
        if(index > rowCnt):
            break

    wb.save(Statics_result_file)

